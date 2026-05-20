# app_jade / CORE / generador_lista_asistencia.py

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from .calendario_asistencia import PeriodoAsistencia


# ============================================================
# CONFIGURACIÓN GENERAL
# ============================================================

DESTINATARIOS_POR_PAGINA_SEMANAL = 38
DESTINATARIOS_POR_PAGINA_MENSUAL = 42


# ============================================================
# ESTILOS BASE
# ============================================================

BORDE_FINO = Side(style="thin", color="000000")

BORDE_COMPLETO = Border(
    left=BORDE_FINO,
    right=BORDE_FINO,
    top=BORDE_FINO,
    bottom=BORDE_FINO,
)

BLANCO = "FFFFFF"
NEGRO = "000000"
AZUL_OSCURO = "173B63"

FUENTE_INFO = Font(name="Arial", size=10, bold=True, color=AZUL_OSCURO)
FUENTE_HEADER = Font(name="Arial", size=8, bold=True, color=NEGRO)
FUENTE_NORMAL = Font(name="Arial", size=12, color=NEGRO)

RELLENO_BLANCO = PatternFill(fill_type="solid", fgColor=BLANCO)


# ============================================================
# HELPERS
# ============================================================

def texto_mayuscula(valor: Any) -> str:
    if valor is None:
        return ""

    return str(valor).strip().upper()


def nombre_completo(destinatario: dict[str, Any]) -> str:
    nombre = texto_mayuscula(destinatario.get("nombre", ""))
    apellido = texto_mayuscula(destinatario.get("apellido", ""))

    return f"{apellido}, {nombre}".strip(", ")


def obtener_texto_periodo(periodo: PeriodoAsistencia) -> str:
    return periodo.etiqueta_periodo


def obtener_label_periodo(periodo: PeriodoAsistencia) -> str:
    if periodo.tipo == "semanal":
        return "SEMANA"

    if periodo.tipo == "mensual":
        return "MES"

    return "PERÍODO"


def obtener_destinatarios_por_pagina(periodo: PeriodoAsistencia) -> int:
    if periodo.tipo == "semanal":
        return DESTINATARIOS_POR_PAGINA_SEMANAL

    return DESTINATARIOS_POR_PAGINA_MENSUAL


def dividir_en_paginas(
    destinatarios: list[dict[str, Any]],
    cantidad_por_pagina: int,
) -> list[list[dict[str, Any]]]:
    if not destinatarios:
        return [[]]

    paginas = []

    for inicio in range(0, len(destinatarios), cantidad_por_pagina):
        fin = inicio + cantidad_por_pagina
        paginas.append(destinatarios[inicio:fin])

    return paginas


def aplicar_estilo_base(
    celda,
    fuente: Font,
    relleno: PatternFill = RELLENO_BLANCO,
    horizontal: str = "center",
    vertical: str = "center",
    wrap_text: bool = True,
) -> None:
    celda.font = fuente
    celda.fill = relleno
    celda.border = BORDE_COMPLETO
    celda.alignment = Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=wrap_text,
    )


def aplicar_estilo_rango(
    ws,
    fila_inicio: int,
    columna_inicio: int,
    fila_fin: int,
    columna_fin: int,
    fuente: Font,
    relleno: PatternFill = RELLENO_BLANCO,
    horizontal: str = "center",
    vertical: str = "center",
) -> None:
    for fila in range(fila_inicio, fila_fin + 1):
        for columna in range(columna_inicio, columna_fin + 1):
            celda = ws.cell(row=fila, column=columna)
            aplicar_estilo_base(
                celda=celda,
                fuente=fuente,
                relleno=relleno,
                horizontal=horizontal,
                vertical=vertical,
            )


# ============================================================
# CONFIGURACIÓN DE PÁGINA
# ============================================================

def configurar_pagina(ws, periodo: PeriodoAsistencia) -> None:
    # Las listas de asistencia se generan en hoja A4 vertical.
    ws.page_setup.orientation = "portrait"

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.30,
        bottom=0.30,
        header=0.10,
        footer=0.10,
    )

    ws.print_options.horizontalCentered = True
    ws.sheet_view.showGridLines = False


def configurar_columnas(ws, cantidad_dias: int) -> None:
    ws.column_dimensions["A"].width = 5

    if cantidad_dias <= 5:
        ws.column_dimensions["B"].width = 34
    else:
        ws.column_dimensions["B"].width = 28

    for indice in range(cantidad_dias):
        letra = get_column_letter(3 + indice)

        if cantidad_dias <= 5:
            ws.column_dimensions[letra].width = 15
        else:
            ws.column_dimensions[letra].width = 6.2


def configurar_alturas_bloque(ws, fila_inicio_bloque: int) -> None:
    ws.row_dimensions[fila_inicio_bloque].height = 20
    ws.row_dimensions[fila_inicio_bloque + 1].height = 20
    ws.row_dimensions[fila_inicio_bloque + 2].height = 20
    ws.row_dimensions[fila_inicio_bloque + 3].height = 20
    ws.row_dimensions[fila_inicio_bloque + 4].height = 34


# ============================================================
# ENCABEZADO UNIFICADO
# ============================================================

def escribir_encabezado_unificado(
    ws,
    sede: str,
    periodo: PeriodoAsistencia,
    ultima_columna: int,
    fila_inicio: int,
    numero_pagina: int,
    total_paginas: int,
) -> None:
    texto_encabezado = (
        f"SEDE: {texto_mayuscula(sede)}\n"
        f"{obtener_label_periodo(periodo)}: {texto_mayuscula(obtener_texto_periodo(periodo))}\n"
        f"FECHA DE EMISIÓN: {periodo.fecha_emision}\n"
        f"PÁGINA: {numero_pagina} DE {total_paginas}"
    )

    fila_fin = fila_inicio + 3

    aplicar_estilo_rango(
        ws=ws,
        fila_inicio=fila_inicio,
        columna_inicio=1,
        fila_fin=fila_fin,
        columna_fin=ultima_columna,
        fuente=FUENTE_INFO,
        relleno=RELLENO_BLANCO,
        horizontal="left",
        vertical="center",
    )

    ws.merge_cells(
        start_row=fila_inicio,
        start_column=1,
        end_row=fila_fin,
        end_column=ultima_columna,
    )

    celda = ws.cell(row=fila_inicio, column=1)
    celda.value = texto_encabezado
    celda.font = FUENTE_INFO
    celda.fill = RELLENO_BLANCO
    celda.border = BORDE_COMPLETO
    celda.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )


# ============================================================
# TABLA DE ASISTENCIA
# ============================================================

def escribir_encabezado_tabla(
    ws,
    periodo: PeriodoAsistencia,
    fila: int,
) -> None:
    encabezados = ["N°", "NOMBRE Y APELLIDO"]

    for dia in periodo.dias:
        if periodo.tipo == "semanal":
            encabezados.append(f"{dia.nombre_dia}\n{dia.fecha_formateada}")
        else:
            encabezados.append(f"{dia.fecha.day}\n{dia.nombre_dia[:3]}")

    for columna, texto in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila, column=columna)
        celda.value = texto
        aplicar_estilo_base(
            celda=celda,
            fuente=FUENTE_HEADER,
            relleno=RELLENO_BLANCO,
            horizontal="center",
        )


def escribir_destinatarios(
    ws,
    destinatarios: list[dict[str, Any]],
    periodo: PeriodoAsistencia,
    fila_inicial: int,
    numero_inicial: int,
) -> int:
    cantidad_columnas = 2 + len(periodo.dias)

    if not destinatarios:
        aplicar_estilo_rango(
            ws=ws,
            fila_inicio=fila_inicial,
            columna_inicio=1,
            fila_fin=fila_inicial,
            columna_fin=cantidad_columnas,
            fuente=FUENTE_NORMAL,
            relleno=RELLENO_BLANCO,
            horizontal="center",
        )

        ws.merge_cells(
            start_row=fila_inicial,
            start_column=1,
            end_row=fila_inicial,
            end_column=cantidad_columnas,
        )

        celda = ws.cell(row=fila_inicial, column=1)
        celda.value = "NO HAY DESTINATARIOS REGISTRADOS PARA ESTE TURNO"
        aplicar_estilo_base(
            celda=celda,
            fuente=FUENTE_NORMAL,
            horizontal="center",
        )

        ws.row_dimensions[fila_inicial].height = 26
        return fila_inicial

    for indice, destinatario in enumerate(destinatarios, start=numero_inicial):
        fila = fila_inicial + (indice - numero_inicial)

        ws.cell(row=fila, column=1).value = indice
        ws.cell(row=fila, column=2).value = nombre_completo(destinatario)

        for columna in range(3, cantidad_columnas + 1):
            ws.cell(row=fila, column=columna).value = ""

        for columna in range(1, cantidad_columnas + 1):
            celda = ws.cell(row=fila, column=columna)
            aplicar_estilo_base(
                celda=celda,
                fuente=FUENTE_NORMAL,
                relleno=RELLENO_BLANCO,
                horizontal="center" if columna != 2 else "left",
            )

        ws.row_dimensions[fila].height = 24

    return fila_inicial + len(destinatarios) - 1


def escribir_bloque_pagina(
    ws,
    destinatarios_pagina: list[dict[str, Any]],
    periodo: PeriodoAsistencia,
    sede: str,
    ultima_columna: int,
    fila_inicio_bloque: int,
    numero_pagina: int,
    total_paginas: int,
    numero_inicial_destinatario: int,
) -> int:
    configurar_alturas_bloque(ws, fila_inicio_bloque)

    escribir_encabezado_unificado(
        ws=ws,
        sede=sede,
        periodo=periodo,
        ultima_columna=ultima_columna,
        fila_inicio=fila_inicio_bloque,
        numero_pagina=numero_pagina,
        total_paginas=total_paginas,
    )

    fila_encabezado_tabla = fila_inicio_bloque + 4
    escribir_encabezado_tabla(
        ws=ws,
        periodo=periodo,
        fila=fila_encabezado_tabla,
    )

    fila_datos = fila_encabezado_tabla + 1

    ultima_fila = escribir_destinatarios(
        ws=ws,
        destinatarios=destinatarios_pagina,
        periodo=periodo,
        fila_inicial=fila_datos,
        numero_inicial=numero_inicial_destinatario,
    )

    return ultima_fila


# ============================================================
# FUNCIÓN PRINCIPAL
# ============================================================

def generar_lista_asistencia(
    destinatarios: list[dict[str, Any]],
    turno: str,
    ruta_salida: str | Path,
    periodo: PeriodoAsistencia,
    sede: str = "",
) -> Path:
    ruta_salida = Path(ruta_salida)
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active

    if periodo.tipo == "semanal":
        ws.title = f"Semanal {turno}"
    else:
        ws.title = f"Mensual {turno}"

    cantidad_dias = len(periodo.dias)
    ultima_columna = 2 + cantidad_dias

    configurar_pagina(ws, periodo)
    configurar_columnas(ws, cantidad_dias)

    destinatarios_por_pagina = obtener_destinatarios_por_pagina(periodo)

    paginas_destinatarios = dividir_en_paginas(
        destinatarios=destinatarios,
        cantidad_por_pagina=destinatarios_por_pagina,
    )

    total_paginas = len(paginas_destinatarios)

    fila_inicio_bloque = 1
    ultima_fila_general = 1

    for indice_pagina, destinatarios_pagina in enumerate(paginas_destinatarios, start=1):
        numero_inicial_destinatario = (
            ((indice_pagina - 1) * destinatarios_por_pagina) + 1
        )

        ultima_fila_bloque = escribir_bloque_pagina(
            ws=ws,
            destinatarios_pagina=destinatarios_pagina,
            periodo=periodo,
            sede=sede,
            ultima_columna=ultima_columna,
            fila_inicio_bloque=fila_inicio_bloque,
            numero_pagina=indice_pagina,
            total_paginas=total_paginas,
            numero_inicial_destinatario=numero_inicial_destinatario,
        )

        ultima_fila_general = ultima_fila_bloque

        if indice_pagina < total_paginas:
            ws.row_breaks.append(Break(id=ultima_fila_bloque))
            fila_inicio_bloque = ultima_fila_bloque + 1

    ultima_letra = get_column_letter(ultima_columna)
    ws.print_area = f"A1:{ultima_letra}{ultima_fila_general}"

    wb.save(ruta_salida)

    return ruta_salida