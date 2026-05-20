# app_jade / CORE / generador.excel.py  
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins


def texto_mayuscula(valor: Any) -> str:
    if valor is None:
        return ""

    return str(valor).strip().upper()


def nombre_completo(nombre: str, apellido: str) -> str:
    nombre = texto_mayuscula(nombre)
    apellido = texto_mayuscula(apellido)

    return f"{nombre} {apellido}".strip()


BORDE_FINO = Side(style="thin", color="000000")
BORDE_GRUESO = Side(style="medium", color="000000")

BORDE_COMPLETO_FINO = Border(
    left=BORDE_FINO,
    right=BORDE_FINO,
    top=BORDE_FINO,
    bottom=BORDE_FINO,
)

FUENTE_NORMAL = Font(name="Arial", size=10)
FUENTE_LABEL = Font(name="Arial", size=10, bold=True, italic=True)
FUENTE_TITULO = Font(name="Arial", size=12, bold=True)
FUENTE_BLOQUE = Font(name="Arial", size=11, bold=True)


def aplicar_borde_rango(ws, rango: str, borde: Border = BORDE_COMPLETO_FINO) -> None:
    for fila in ws[rango]:
        for celda in fila:
            celda.border = borde


def escribir_celda(
    ws,
    celda: str,
    valor: Any,
    fuente: Font = FUENTE_NORMAL,
    alineacion: Alignment | None = None,
    borde: bool = False,
) -> None:
    ws[celda] = texto_mayuscula(valor)
    ws[celda].font = fuente

    if alineacion is None:
        ws[celda].alignment = Alignment(vertical="center")
    else:
        ws[celda].alignment = alineacion

    if borde:
        ws[celda].border = BORDE_COMPLETO_FINO


def escribir_label(ws, celda: str, texto: str) -> None:
    escribir_celda(
        ws=ws,
        celda=celda,
        valor=texto,
        fuente=FUENTE_LABEL,
        alineacion=Alignment(horizontal="left", vertical="center"),
        borde=True,
    )


def escribir_valor(ws, celda: str, valor: Any) -> None:
    escribir_celda(
        ws=ws,
        celda=celda,
        valor=valor,
        fuente=FUENTE_NORMAL,
        alineacion=Alignment(horizontal="center", vertical="center"),
        borde=True,
    )


def unir_y_estilizar(ws, rango: str, valor: Any = "", fuente: Font = FUENTE_NORMAL) -> None:
    ws.merge_cells(rango)
    celda_inicio = rango.split(":")[0]

    ws[celda_inicio] = texto_mayuscula(valor)
    ws[celda_inicio].font = fuente
    ws[celda_inicio].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    aplicar_borde_rango(ws, rango)


def unir_label(ws, rango: str, texto: str) -> None:
    ws.merge_cells(rango)
    celda_inicio = rango.split(":")[0]

    ws[celda_inicio] = texto_mayuscula(texto)
    ws[celda_inicio].font = FUENTE_LABEL
    ws[celda_inicio].alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    aplicar_borde_rango(ws, rango)


def generar_solicitud_alta(datos: dict[str, Any], ruta_salida: str | Path) -> Path:
    ruta_salida = Path(ruta_salida)
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitud de alta"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.35,
        bottom=0.35,
        header=0.1,
        footer=0.1,
    )

    anchos = {
        "A": 3,
        "B": 15,
        "C": 15,
        "D": 15,
        "E": 15,
        "F": 15,
        "G": 15,
        "H": 3,
        "I": 10,
        "J": 10,
        "K": 10,
        "L": 10,
        "M": 10,
        "N": 10,
        "O": 3,
    }

    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    for fila in range(1, 55):
        ws.row_dimensions[fila].height = 20

    ws.row_dimensions[5].height = 24

    for fila in range(21, 36):
        ws.row_dimensions[fila].height = 24

    fecha = datos.get("fecha", "")
    origen = datos.get("origen", "")
    municipio = datos.get("municipio", "")
    sede = datos.get("sede", "")
    barrio = datos.get("barrio", "")
    localidad = datos.get("localidad", "")

    destinatario_nombre_apellido = nombre_completo(
        datos.get("destinatario_nombre", ""),
        datos.get("destinatario_apellido", ""),
    )
    destinatario_dni = datos.get("destinatario_dni", "")
    destinatario_direccion = datos.get("destinatario_direccion", "")
    destinatario_fecha_nacimiento = datos.get("destinatario_fecha_nacimiento", "")
    destinatario_edad = datos.get("destinatario_edad", "")
    destinatario_escolarizado = datos.get("destinatario_escolarizado", "")

    responsable_nombre_apellido = nombre_completo(
        datos.get("responsable_nombre", ""),
        datos.get("responsable_apellido", ""),
    )
    responsable_telefono = datos.get("responsable_telefono", "")
    responsable_domicilio = datos.get("responsable_domicilio", "")
    responsable_dni = datos.get("responsable_dni", "")
    responsable_parentesco = datos.get("responsable_parentesco", "")
    responsable_fecha_nacimiento = datos.get("responsable_fecha_nacimiento", "")
    responsable_edad = datos.get("responsable_edad", "")

    ws.merge_cells("C5:G5")
    ws["C5"] = "SOLICITUD DE ASPIRANTE"
    ws["C5"].font = FUENTE_TITULO
    ws["C5"].alignment = Alignment(horizontal="center", vertical="center")

    escribir_label(ws, "B7", "FECHA:")
    unir_y_estilizar(ws, "C7:D7", fecha)

    escribir_label(ws, "E7", "ORIGEN:")
    unir_y_estilizar(ws, "F7:G7", origen)

    escribir_label(ws, "B9", "MUNICIPIO:")
    unir_y_estilizar(ws, "C9:D9", municipio)

    escribir_label(ws, "E9", "SEDE:")
    unir_y_estilizar(ws, "F9:G9", sede)

    unir_label(ws, "B11:C11", "NOMBRE Y APELLIDO:")
    unir_y_estilizar(ws, "D11:G11", destinatario_nombre_apellido)

    escribir_label(ws, "B13", "DNI:")
    unir_y_estilizar(ws, "C13:D13", destinatario_dni)

    escribir_label(ws, "E13", "DIRECCIÓN:")
    unir_y_estilizar(ws, "F13:G13", destinatario_direccion)

    escribir_label(ws, "B15", "BARRIO:")
    unir_y_estilizar(ws, "C15:D15", barrio)

    escribir_label(ws, "E15", "LOCALIDAD:")
    unir_y_estilizar(ws, "F15:G15", localidad)

    unir_label(ws, "B17:C17", "FECHA DE NACIMIENTO:")
    escribir_valor(ws, "D17", destinatario_fecha_nacimiento)

    escribir_label(ws, "E17", "EDAD:")
    unir_y_estilizar(ws, "F17:G17", destinatario_edad)

    unir_label(ws, "B19:C19", "ESCOLARIZADO:")
    unir_y_estilizar(ws, "D19:G19", destinatario_escolarizado)

    ws.merge_cells("B21:G35")
    ws["B21"] = (
        "EL EQUIPO INTERDISCIPLINARIO DEL PROGRAMA DE RESPONSABILIDAD "
        "SOCIAL COMPARTIDA ENVIÓN"
    )
    ws["B21"].font = FUENTE_NORMAL
    ws["B21"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    for fila in range(21, 36):
        for col in range(2, 8):
            celda = ws.cell(row=fila, column=col)
            celda.border = Border(
                left=BORDE_GRUESO if col == 2 else Side(style=None),
                right=BORDE_GRUESO if col == 7 else Side(style=None),
                top=BORDE_GRUESO if fila == 21 else Side(style=None),
                bottom=BORDE_GRUESO if fila == 35 else Side(style=None),
            )

    ws.merge_cells("C37:G37")
    ws["C37"] = "DATOS RESPONSABLE ADULTO"
    ws["C37"].font = FUENTE_TITULO
    ws["C37"].alignment = Alignment(horizontal="center", vertical="center")

    unir_label(ws, "B39:C39", "NOMBRE Y APELLIDO:")
    unir_y_estilizar(ws, "D39:G39", responsable_nombre_apellido)

    unir_label(ws, "B41:C41", "TEL DE CONTACTO:")
    unir_y_estilizar(ws, "D41:G41", responsable_telefono)

    unir_label(ws, "B43:C43", "DOMICILIO:")
    unir_y_estilizar(ws, "D43:G43", responsable_domicilio)

    escribir_label(ws, "B45", "DNI:")
    unir_y_estilizar(ws, "C45:D45", responsable_dni)

    escribir_label(ws, "E45", "PARENTESCO:")
    unir_y_estilizar(ws, "F45:G45", responsable_parentesco)

    unir_label(ws, "B47:C47", "FECHA DE NACIMIENTO:")
    escribir_valor(ws, "D47", responsable_fecha_nacimiento)

    escribir_label(ws, "E47", "EDAD:")
    unir_y_estilizar(ws, "F47:G47", responsable_edad)

    ws.merge_cells("D53:E53")
    ws["D53"] = "21/04/2026 REV. 00"
    ws["D53"].font = Font(name="Arial", size=9)
    ws["D53"].alignment = Alignment(horizontal="center", vertical="center")

    ws.print_area = "A1:G54"

    wb.save(ruta_salida)

    return ruta_salida